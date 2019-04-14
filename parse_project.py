import openpyxl

from attr import attrs, attrib
from collections import defaultdict


@attrs
class Rig:
    name = attrib(default='')
    is_needed = attrib(default=False)
    board_name = attrib(default='')
    test_date = attrib(default='')
    dev_name = attrib(default='')
    is_received = attrib(default=False)


wb = openpyxl.load_workbook('.\\ref\\in.xlsx')
ws = wb.active

b_col = 2

dev_ranges = list()
for rng in ws.merged_cells:
    if rng.min_col == b_col:
        dev_ranges.append(rng)

dev_ranges = list(reversed(dev_ranges))

devs = defaultdict(list)
for dev in dev_ranges:
    col, min_row, _, max_row = dev.bounds
    dev_name = ws.cell(min_row, col).value
    print(dev_name)
    for row in range(min_row, max_row + 1):
        r = Rig(
            name=ws.cell(row, col + 1).value,
            is_needed=True if ws.cell(row, col + 2).value == '+' else False,
            board_name=ws.cell(row, col + 3).value,
            test_date=ws.cell(row, col + 4).value,
            dev_name=ws.cell(row, col + 5).value,
            is_received=ws.cell(row, col + 6).value
        )
        print(r)
        devs[dev_name].append(r)

print(devs)

wb.close()
